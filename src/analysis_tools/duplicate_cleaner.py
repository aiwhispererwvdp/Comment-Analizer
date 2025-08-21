"""
Duplicate Detection and Cleaning Tool
Identifies and removes duplicate comments while tracking frequency
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Tuple, Optional
from collections import Counter
import hashlib
import re
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class DuplicateCleaner:
    """Tool for detecting and cleaning duplicate comments with frequency tracking"""
    
    def __init__(self, similarity_threshold: float = 0.95):
        """
        Initialize duplicate cleaner
        
        Args:
            similarity_threshold: Threshold for fuzzy matching (0-1)
        """
        self.similarity_threshold = similarity_threshold
        self.duplicate_stats = {}
        self.cleaned_data = None
        
    def clean_text(self, text: str) -> str:
        """Normalize text for comparison"""
        if pd.isna(text):
            return ""
        
        # Convert to lowercase and remove extra spaces
        text = str(text).lower().strip()
        text = re.sub(r'\s+', ' ', text)
        
        # Remove punctuation for comparison
        text = re.sub(r'[^\w\s]', '', text)
        
        return text
    
    def generate_hash(self, text: str) -> str:
        """Generate hash for exact duplicate detection"""
        cleaned = self.clean_text(text)
        return hashlib.md5(cleaned.encode()).hexdigest()
    
    def find_exact_duplicates(self, df: pd.DataFrame, text_column: str = 'Comentario Final') -> Dict:
        """
        Find exact duplicates in dataset
        
        Args:
            df: DataFrame with comments
            text_column: Column name containing text
            
        Returns:
            Dictionary with duplicate analysis
        """
        # Create hash column
        df['text_hash'] = df[text_column].apply(self.generate_hash)
        
        # Count occurrences
        hash_counts = df['text_hash'].value_counts()
        duplicate_hashes = hash_counts[hash_counts > 1]
        
        # Build duplicate groups
        duplicate_groups = {}
        for hash_val, count in duplicate_hashes.items():
            mask = df['text_hash'] == hash_val
            duplicate_rows = df[mask]
            
            # Get original text (first occurrence)
            original_text = duplicate_rows.iloc[0][text_column]
            
            duplicate_groups[hash_val] = {
                'original_text': original_text,
                'count': count,
                'indices': duplicate_rows.index.tolist(),
                'dates': duplicate_rows['Fecha'].tolist() if 'Fecha' in df.columns else None,
                'scores': duplicate_rows['Nota'].tolist() if 'Nota' in df.columns else None
            }
        
        return duplicate_groups
    
    def find_similar_comments(self, df: pd.DataFrame, text_column: str = 'Comentario Final') -> List[Tuple]:
        """
        Find similar (not exact) duplicates using fuzzy matching
        
        Args:
            df: DataFrame with comments
            text_column: Column name containing text
            
        Returns:
            List of similar comment pairs
        """
        from difflib import SequenceMatcher
        
        similar_pairs = []
        texts = df[text_column].dropna().tolist()
        
        for i in range(len(texts)):
            for j in range(i + 1, len(texts)):
                similarity = SequenceMatcher(None, 
                                           self.clean_text(texts[i]), 
                                           self.clean_text(texts[j])).ratio()
                
                if similarity >= self.similarity_threshold:
                    similar_pairs.append({
                        'index1': i,
                        'index2': j,
                        'text1': texts[i][:100] + '...' if len(texts[i]) > 100 else texts[i],
                        'text2': texts[j][:100] + '...' if len(texts[j]) > 100 else texts[j],
                        'similarity': similarity
                    })
        
        return similar_pairs
    
    def remove_duplicates(self, df: pd.DataFrame, 
                         text_column: str = 'Comentario Final',
                         keep: str = 'first',
                         track_frequency: bool = True) -> pd.DataFrame:
        """
        Remove duplicates and track frequency
        
        Args:
            df: DataFrame with comments
            text_column: Column name containing text
            keep: Which duplicate to keep ('first', 'last', 'highest_score')
            track_frequency: Whether to add frequency column
            
        Returns:
            Cleaned DataFrame without duplicates
        """
        df = df.copy()
        
        # Find exact duplicates
        duplicate_groups = self.find_exact_duplicates(df, text_column)
        
        # Add frequency column if requested
        if track_frequency:
            df['frequency'] = 1
            for hash_val, group in duplicate_groups.items():
                df.loc[df['text_hash'] == hash_val, 'frequency'] = group['count']
        
        # Determine which duplicates to keep
        if keep == 'highest_score' and 'Nota' in df.columns:
            # Keep the duplicate with highest score
            indices_to_drop = []
            for hash_val, group in duplicate_groups.items():
                if group['scores']:
                    scores_with_indices = list(zip(group['scores'], group['indices']))
                    scores_with_indices.sort(key=lambda x: x[0], reverse=True)
                    # Drop all except the highest score
                    indices_to_drop.extend([idx for _, idx in scores_with_indices[1:]])
            
            df_cleaned = df.drop(indices_to_drop)
        else:
            # Use pandas built-in drop_duplicates
            df_cleaned = df.drop_duplicates(subset=['text_hash'], keep=keep)
        
        # Remove helper column
        df_cleaned = df_cleaned.drop('text_hash', axis=1)
        
        self.cleaned_data = df_cleaned
        return df_cleaned
    
    def generate_duplicate_report(self, df: pd.DataFrame, 
                                 text_column: str = 'Comentario Final') -> Dict:
        """
        Generate comprehensive duplicate analysis report
        
        Args:
            df: DataFrame with comments
            text_column: Column name containing text
            
        Returns:
            Dictionary with duplicate statistics
        """
        total_comments = len(df)
        duplicate_groups = self.find_exact_duplicates(df, text_column)
        
        # Calculate statistics
        total_duplicates = sum(group['count'] - 1 for group in duplicate_groups.values())
        unique_comments = total_comments - total_duplicates
        duplication_rate = (total_duplicates / total_comments * 100) if total_comments > 0 else 0
        
        # Find most repeated comments
        most_repeated = sorted(duplicate_groups.items(), 
                              key=lambda x: x[1]['count'], 
                              reverse=True)[:10]
        
        # Analyze duplicate patterns
        duplicate_scores = []
        for group in duplicate_groups.values():
            if group['scores']:
                duplicate_scores.extend(group['scores'])
        
        avg_duplicate_score = np.mean(duplicate_scores) if duplicate_scores else 0
        
        report = {
            'summary': {
                'total_comments': total_comments,
                'unique_comments': unique_comments,
                'total_duplicates': total_duplicates,
                'duplication_rate': round(duplication_rate, 2),
                'duplicate_groups': len(duplicate_groups)
            },
            'most_repeated': [
                {
                    'text': group['original_text'][:200] + '...' if len(group['original_text']) > 200 else group['original_text'],
                    'frequency': group['count'],
                    'avg_score': np.mean(group['scores']) if group['scores'] else None
                }
                for _, group in most_repeated
            ],
            'score_analysis': {
                'avg_duplicate_score': round(avg_duplicate_score, 2),
                'score_distribution': Counter(duplicate_scores) if duplicate_scores else {}
            },
            'duplicate_details': duplicate_groups
        }
        
        self.duplicate_stats = report
        return report
    
    def export_duplicate_analysis(self, df: pd.DataFrame, 
                                 output_path: str,
                                 text_column: str = 'Comentario Final'):
        """
        Export duplicate analysis to Excel
        
        Args:
            df: DataFrame with comments
            output_path: Path for output Excel file
            text_column: Column name containing text
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        report = self.generate_duplicate_report(df, text_column)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_df = pd.DataFrame([report['summary']])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Most repeated comments
            repeated_df = pd.DataFrame(report['most_repeated'])
            repeated_df.to_excel(writer, sheet_name='Most Repeated', index=False)
            
            # Cleaned data
            if self.cleaned_data is not None:
                self.cleaned_data.to_excel(writer, sheet_name='Cleaned Data', index=False)
            
            # Format sheets
            workbook = writer.book
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Header formatting
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", 
                                          end_color="366092", 
                                          fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    sheet.column_dimensions[column_letter].width = adjusted_width
        
        logger.info(f"Duplicate analysis exported to {output_path}")